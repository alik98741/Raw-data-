from __future__ import annotations
import json, re, unicodedata
from pathlib import Path
import requests
import pandas as pd
from openpyxl import load_workbook

OUT=Path('sugarcane-industrial-algorithms/mapbiomas-coverage-v1'); OUT.mkdir(parents=True,exist_ok=True)
PID='doi:10.58053/MapBiomas/SJZOLT'
BASE='https://data.mapbiomas.org'
S=requests.Session(); S.headers.update({'User-Agent':'Sugarcane-Industrial-Algorithms/1.0 research'})

def norm(x):
    s=unicodedata.normalize('NFKD',str(x or '')).encode('ascii','ignore').decode('ascii').upper()
    return re.sub(r'[^A-Z0-9]+',' ',s).strip()

def get(url,**kw):
    r=S.get(url,timeout=180,allow_redirects=True,**kw); r.raise_for_status(); return r

meta_url=f'{BASE}/api/datasets/:persistentId/?persistentId={PID}'
r=get(meta_url); meta=r.json()
if meta.get('status')!='OK': raise RuntimeError('Dataverse metadata failed: '+str(meta)[:500])
files=meta['data']['latestVersion']['files']
selected=None
for f in files:
    fn=f['dataFile']['filename']
    if fn.lower().endswith('.xlsx') and 'COVERAGE' in fn.upper(): selected=f['dataFile']; break
if selected is None: raise RuntimeError('Expected MapBiomas coverage xlsx not found: '+str([f['dataFile']['filename'] for f in files]))
file_id=selected['id']; fn=selected['filename']
xlsx=OUT/fn
with get(f'{BASE}/api/access/datafile/{file_id}',stream=True) as rr:
    with xlsx.open('wb') as w:
        for chunk in rr.iter_content(1024*1024):
            if chunk: w.write(chunk)
print('downloaded',fn,xlsx.stat().st_size,flush=True)

# Stream workbook and extract every row containing the sugarcane class. Retain all contextual columns and 1985-2024.
wb=load_workbook(xlsx,read_only=True,data_only=True)
extract=[]; sheet_audit=[]
for ws in wb.worksheets:
    header=None; header_row=None
    # Find a header row containing multiple target years.
    for ridx,row in enumerate(ws.iter_rows(min_row=1,max_row=40,values_only=True),start=1):
        vals=[norm(v) for v in row]
        if '1985' in vals and ('2024' in vals or '2023' in vals):
            header=[str(v).strip() if v is not None else f'COL_{i+1}' for i,v in enumerate(row)]
            header_row=ridx; break
    if header is None:
        sheet_audit.append({'sheet':ws.title,'header_row':None,'sugarcane_rows':0}); continue
    count=0
    for row in ws.iter_rows(min_row=header_row+1,values_only=True):
        nvals=[norm(v) for v in row]
        has_cane=any(('SUGAR CANE' in v) or ('SUGARCANE' in v) or ('CANA DE ACUCAR' in v) for v in nvals)
        if not has_cane: continue
        d={header[i]:row[i] if i<len(row) else None for i in range(len(header))}
        d['_sheet']=ws.title; extract.append(d); count+=1
    sheet_audit.append({'sheet':ws.title,'header_row':header_row,'sugarcane_rows':count})
wb.close()
if not extract:
    raise RuntimeError('No sugarcane rows found. Sheet audit: '+json.dumps(sheet_audit))
raw=pd.DataFrame(extract)
raw.to_csv(OUT/'MapBiomas_Collection10_1_sugarcane_rows_raw.csv',index=False)

# Convert year columns to a long table while keeping all contextual columns.
yearcols=[c for c in raw.columns if re.fullmatch(r'20\d\d|19\d\d',str(c)) and 1985<=int(c)<=2024]
idcols=[c for c in raw.columns if c not in yearcols]
long=raw.melt(id_vars=idcols,value_vars=yearcols,var_name='year',value_name='area_ha')
long['year']=pd.to_numeric(long.year,errors='coerce').astype('Int64')
long['area_ha']=pd.to_numeric(long.area_ha,errors='coerce')
long.to_csv(OUT/'MapBiomas_Collection10_1_sugarcane_1985_2024_long.csv',index=False)

qc={'persistent_id':PID,'file_id':file_id,'filename':fn,'size_bytes':xlsx.stat().st_size,'sheets':sheet_audit,'matched_rows_wide':len(raw),'year_columns':yearcols,'long_rows':len(long),'area_nonmissing':int(long.area_ha.notna().sum()),'status':'PASS'}
(OUT/'MAPBIOMAS_COVERAGE_QC.json').write_text(json.dumps(qc,indent=2,ensure_ascii=False),encoding='utf-8')
print(json.dumps(qc,indent=2,ensure_ascii=False),flush=True)
# Remove 75 MB source from artifact; DOI + file id + MD5 in Dataverse preserve provenance.
xlsx.unlink()
